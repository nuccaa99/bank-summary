import os
import glob
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def find_transactions_sheet(file_path):
    xl = pd.ExcelFile(file_path)
    for sheet in xl.sheet_names:
        if 'transact' in sheet.lower():
            return sheet
    # fallback: second sheet if exists, else first
    return xl.sheet_names[1] if len(xl.sheet_names) >= 2 else xl.sheet_names[0]


def process_tbc(file_path, keyword):
    sheet_name = find_transactions_sheet(file_path)
    df = pd.read_excel(file_path, sheet_name=sheet_name)

    keyword_nospace = keyword.replace(' ', '')
    mask = df['Description'].astype(str).apply(
        lambda x: keyword_nospace.lower() in x.replace(' ', '').lower()
    )
    filtered = df[mask].copy()

    if filtered.empty:
        return None

    filtered['Paid Out'] = pd.to_numeric(filtered['Paid Out'], errors='coerce').fillna(0)

    result = filtered.groupby("Partner's Name")['Paid Out'].sum().reset_index()
    result.columns = ['Partner Name', 'Total Amount']
    result = result[result['Total Amount'] != 0]
    if result.empty:
        return None
    result['Bank'] = 'TBC'
    result['Source File'] = os.path.basename(file_path)
    return result


def process_bog(file_path, keyword):
    df = pd.read_excel(file_path, sheet_name=0)

    keyword_nospace = keyword.replace(' ', '')
    mask = df['დანიშნულება'].astype(str).apply(
        lambda x: keyword_nospace.lower() in x.replace(' ', '').lower()
    )
    filtered = df[mask].copy()

    if filtered.empty:
        return None

    filtered['დებეტი'] = pd.to_numeric(filtered['დებეტი'], errors='coerce').fillna(0)

    result = filtered.groupby('მიმღების დასახელება')['დებეტი'].sum().reset_index()
    result.columns = ['Partner Name', 'Total Amount']
    result = result[result['Total Amount'] != 0]
    if result.empty:
        return None
    result['Bank'] = 'BOG'
    result['Source File'] = os.path.basename(file_path)
    return result


def process_tbc_credits(file_path):
    sheet_name = find_transactions_sheet(file_path)
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    df['Paid In'] = pd.to_numeric(df['Paid In'], errors='coerce').fillna(0)
    result = df.groupby("Partner's Name")['Paid In'].sum().reset_index()
    result.columns = ['Partner Name', 'Total Amount']
    result = result[result['Total Amount'] != 0]
    if result.empty:
        return None
    result['Bank'] = 'TBC'
    result['Source File'] = os.path.basename(file_path)
    return result


def process_bog_credits(file_path):
    df = pd.read_excel(file_path, sheet_name=0)
    df['კრედიტი'] = pd.to_numeric(df['კრედიტი'], errors='coerce').fillna(0)
    result = df.groupby('მიმღების დასახელება')['კრედიტი'].sum().reset_index()
    result.columns = ['Partner Name', 'Total Amount']
    result = result[result['Total Amount'] != 0]
    if result.empty:
        return None
    result['Bank'] = 'BOG'
    result['Source File'] = os.path.basename(file_path)
    return result


def build_credits_df(tbc_files, bog_files, name_filter):
    rows = []
    for f in tbc_files:
        if name_filter.lower() in os.path.basename(f).lower():
            try:
                r = process_tbc_credits(f)
                if r is not None:
                    rows.append(r)
            except Exception as e:
                print(f"  Credits TBC error ({os.path.basename(f)}): {e}")
    for f in bog_files:
        if name_filter.lower() in os.path.basename(f).lower():
            try:
                r = process_bog_credits(f)
                if r is not None:
                    rows.append(r)
            except Exception as e:
                print(f"  Credits BOG error ({os.path.basename(f)}): {e}")
    if not rows:
        return None
    combined = pd.concat(rows, ignore_index=True)
    grand_total = combined['Total Amount'].sum()
    return pd.DataFrame([{'Bank': '', 'Partner Name': '--- TOTAL ---', 'Total Amount': grand_total, 'Source File': ''}])


def process_tbc_salary(file_path):
    sheet_name = find_transactions_sheet(file_path)
    df = pd.read_excel(file_path, sheet_name=sheet_name)
    mask = df['Description'].astype(str).apply(
        lambda x: 'ხელფასი' in x.replace(' ', '').lower() or 'ხელფასი' in x.lower()
    )
    filtered = df[mask].copy()
    if filtered.empty:
        return None
    filtered['Paid Out'] = pd.to_numeric(filtered['Paid Out'], errors='coerce').fillna(0)
    filtered = filtered[filtered['Paid Out'] != 0]
    if filtered.empty:
        return None
    result = filtered[['Date', "Partner's Name", 'Paid Out']].copy()
    result.columns = ['Date', 'Partner Name', 'Amount']
    result['Bank'] = 'TBC'
    result['Source File'] = os.path.basename(file_path)
    return result


def process_bog_salary(file_path):
    df = pd.read_excel(file_path, sheet_name=0)
    mask = df['დანიშნულება'].astype(str).apply(
        lambda x: 'ხელფასი' in x.replace(' ', '').lower() or 'ხელფასი' in x.lower()
    )
    filtered = df[mask].copy()
    if filtered.empty:
        return None
    filtered['დებეტი'] = pd.to_numeric(filtered['დებეტი'], errors='coerce').fillna(0)
    filtered = filtered[filtered['დებეტი'] != 0]
    if filtered.empty:
        return None
    result = filtered[['თარიღი', 'მიმღების დასახელება', 'დებეტი']].copy()
    result.columns = ['Date', 'Partner Name', 'Amount']
    result['Bank'] = 'BOG'
    result['Source File'] = os.path.basename(file_path)
    return result


def main():
    keyword = input("Enter keyword to filter descriptions by: ").strip()
    if not keyword:
        print("No keyword entered. Exiting.")
        return

    script_dir = os.path.dirname(os.path.abspath(__file__))

    tbc_files = glob.glob(os.path.join(script_dir, '*TBC*'))
    bog_files = glob.glob(os.path.join(script_dir, '*BOG*'))

    if not tbc_files and not bog_files:
        print("No TBC or BOG files found in the script folder.")
        return

    results = []

    for f in tbc_files:
        print(f"Processing TBC: {os.path.basename(f)}")
        try:
            r = process_tbc(f, keyword)
            if r is not None:
                results.append(r)
            else:
                print(f"  No matching rows found.")
        except Exception as e:
            print(f"  Error: {e}")

    for f in bog_files:
        print(f"Processing BOG: {os.path.basename(f)}")
        try:
            r = process_bog(f, keyword)
            if r is not None:
                results.append(r)
            else:
                print(f"  No matching rows found.")
        except Exception as e:
            print(f"  Error: {e}")

    if not results:
        print("\nNo matching data found for this keyword.")
        input("Press Enter to exit.")
        return

    combined = pd.concat(results, ignore_index=True)
    combined = combined[['Bank', 'Partner Name', 'Total Amount', 'Source File']]

    # Normalize names: sort words so "John Smith" and "Smith John" are treated as the same person
    def normalize_name(name):
        return ' '.join(sorted(str(name).split()))

    combined['_name_key'] = combined['Partner Name'].apply(normalize_name)

    # Pick the canonical name (first occurrence) and sum amounts per normalized key + bank
    canonical_names = combined.groupby('_name_key')['Partner Name'].first()
    combined['Partner Name'] = combined['_name_key'].map(canonical_names)
    combined = combined.groupby(['Bank', 'Partner Name', '_name_key', 'Source File'], as_index=False).agg(
        {'Total Amount': 'sum'}
    )
    combined = combined.drop(columns=['_name_key'])

    combined = combined.sort_values(['Source File', 'Total Amount'], ascending=[True, False])

    # Add totals per source file
    rows = []
    for source_file, group in combined.groupby('Source File', sort=True):
        rows.append(group)
        total_row = pd.DataFrame([{
            'Bank': '',
            'Partner Name': f'--- {source_file} TOTAL ---',
            'Total Amount': group['Total Amount'].sum(),
            'Source File': source_file
        }])
        rows.append(total_row)

    final = pd.concat(rows, ignore_index=True)

    # Build combined summary: merge same person across all banks
    summary = combined.copy()
    summary['_name_key'] = summary['Partner Name'].apply(lambda n: ' '.join(sorted(str(n).split())))
    canonical = summary.groupby('_name_key')['Partner Name'].first()
    summary['Partner Name'] = summary['_name_key'].map(canonical)
    summary = summary.groupby(['Partner Name', '_name_key'], as_index=False).agg(
        {'Total Amount': 'sum',
         'Bank': lambda x: ', '.join(sorted(set(x))),
         'Source File': lambda x: ', '.join(sorted(set(x)))}
    )
    summary = summary.drop(columns=['_name_key'])
    summary = summary[['Bank', 'Partner Name', 'Total Amount', 'Source File']]
    summary = summary.sort_values('Total Amount', ascending=False)
    grand_total = pd.DataFrame([{
        'Bank': '', 'Partner Name': '--- GRAND TOTAL ---',
        'Total Amount': summary['Total Amount'].sum(), 'Source File': ''
    }])
    summary = pd.concat([summary, grand_total], ignore_index=True)

    # Build company-specific summary sheets
    def build_company_df(combined, filter_str):
        subset = combined[combined['Source File'].str.contains(filter_str, case=False, na=False)].copy()
        if subset.empty:
            return None
        subset['_name_key'] = subset['Partner Name'].apply(lambda n: ' '.join(sorted(str(n).split())))
        canonical = subset.groupby('_name_key')['Partner Name'].first()
        subset['Partner Name'] = subset['_name_key'].map(canonical)
        grouped = subset.groupby(['Partner Name', '_name_key'], as_index=False).agg(
            {'Total Amount': 'sum',
             'Bank': lambda x: ', '.join(sorted(set(x))),
             'Source File': lambda x: ', '.join(sorted(set(x)))}
        ).drop(columns=['_name_key'])
        grouped = grouped[['Bank', 'Partner Name', 'Total Amount', 'Source File']]
        grouped = grouped[grouped['Total Amount'] != 0]
        grouped = grouped.sort_values('Total Amount', ascending=False)
        total = pd.DataFrame([{
            'Bank': '', 'Partner Name': '--- TOTAL ---',
            'Total Amount': grouped['Total Amount'].sum(), 'Source File': ''
        }])
        return pd.concat([grouped, total], ignore_index=True)

    del_df  = build_company_df(combined, 'დელ')
    cel_df  = build_company_df(combined, 'ცელ')

    # Build credit sheets for ზუკა and ირმა files
    zuka_df = build_credits_df(tbc_files, bog_files, 'ზუკა')
    irma_df = build_credits_df(tbc_files, bog_files, 'ირმა')

    # Build salary sheet: individual transactions with dates for "ხელფასი"
    salary_rows = []
    for f in tbc_files:
        try:
            r = process_tbc_salary(f)
            if r is not None:
                salary_rows.append(r)
        except Exception as e:
            print(f"  Salary TBC error ({os.path.basename(f)}): {e}")
    for f in bog_files:
        try:
            r = process_bog_salary(f)
            if r is not None:
                salary_rows.append(r)
        except Exception as e:
            print(f"  Salary BOG error ({os.path.basename(f)}): {e}")

    if salary_rows:
        salary_df = pd.concat(salary_rows, ignore_index=True)
        salary_df = salary_df[['Date', 'Bank', 'Partner Name', 'Amount', 'Source File']]
        salary_df = salary_df.sort_values('Source File')
        salary_total = pd.DataFrame([{
            'Date': '', 'Bank': '', 'Partner Name': '--- TOTAL ---',
            'Amount': salary_df['Amount'].sum(), 'Source File': ''
        }])
        salary_df = pd.concat([salary_df, salary_total], ignore_index=True)
    else:
        salary_df = None

    output_path = os.path.join(script_dir, 'output.xlsx')

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        final.to_excel(writer, index=False, sheet_name='Results')
        summary.to_excel(writer, index=False, sheet_name='Summary')
        if salary_df is not None:
            salary_df.to_excel(writer, index=False, sheet_name='Salary')
        if del_df is not None:
            del_df.to_excel(writer, index=False, sheet_name='მომსახ დელ')
        if cel_df is not None:
            cel_df.to_excel(writer, index=False, sheet_name='მომსახ ცელ')
        if zuka_df is not None:
            zuka_df.to_excel(writer, index=False, sheet_name='ზუკა')
        if irma_df is not None:
            irma_df.to_excel(writer, index=False, sheet_name='ირმა')
        ws = writer.sheets['Results']

        # Header style
        header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill

        # Row colors
        tbc_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        bog_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        total_fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')

        # Assign a unique color per source file
        file_color_palette = [
            'D9E1F2', 'E2EFDA', 'FCE4D6', 'EAD1DC', 'D0E4F7',
            'FFF2CC', 'E8D5F0', 'D5F0E8', 'F0E8D5', 'D5E8F0',
        ]
        unique_files = final['Source File'].dropna().unique()
        unique_files = [f for f in unique_files if f != '']
        file_color_map = {
            f: file_color_palette[i % len(file_color_palette)]
            for i, f in enumerate(sorted(unique_files))
        }

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            partner = str(row[1].value or '')
            source = str(row[3].value or '')
            is_total = 'TOTAL' in partner
            if is_total:
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = total_fill
            else:
                color = file_color_map.get(source, 'F2F2F2')
                fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                for cell in row:
                    cell.fill = fill

        # Auto-fit columns
        for col in ws.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

        # Style Summary sheet
        ws2 = writer.sheets['Summary']
        for cell in ws2[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = header_fill
        summary_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        for row in ws2.iter_rows(min_row=2, max_row=ws2.max_row):
            partner = str(row[1].value or '')
            is_total = 'TOTAL' in partner
            for cell in row:
                if is_total:
                    cell.font = Font(bold=True)
                    cell.fill = total_fill
                else:
                    cell.fill = summary_fill
        for col in ws2.columns:
            max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
            ws2.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

        # Style Salary sheet
        if salary_df is not None:
            ws3 = writer.sheets['Salary']
            for cell in ws3[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = header_fill
            for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
                partner = str(row[2].value or '')
                bank = str(row[1].value or '')
                is_total = 'TOTAL' in partner
                for cell in row:
                    if is_total:
                        cell.font = Font(bold=True)
                        cell.fill = total_fill
                    else:
                        cell.fill = tbc_fill if bank == 'TBC' else bog_fill
            for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
                if row[0].value and str(row[0].value) != '':
                    row[0].number_format = 'DD/MM/YYYY'
            for col in ws3.columns:
                max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
                ws3.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

        # Style company sheets
        for sheet_name, df in [('მომსახ დელ', del_df), ('მომსახ ცელ', cel_df), ('ზუკა', zuka_df), ('ირმა', irma_df)]:
            if df is None:
                continue
            ws_c = writer.sheets[sheet_name]
            for cell in ws_c[1]:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = header_fill
            for row in ws_c.iter_rows(min_row=2, max_row=ws_c.max_row):
                partner = str(row[1].value or '')
                is_total = 'TOTAL' in partner
                for cell in row:
                    if is_total:
                        cell.font = Font(bold=True)
                        cell.fill = total_fill
                    else:
                        cell.fill = summary_fill
            for col in ws_c.columns:
                max_len = max((len(str(cell.value)) if cell.value else 0) for cell in col)
                ws_c.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

    print(f"\nDone! Output saved to: {output_path}")
    input("Press Enter to exit.")


if __name__ == '__main__':
    main()
